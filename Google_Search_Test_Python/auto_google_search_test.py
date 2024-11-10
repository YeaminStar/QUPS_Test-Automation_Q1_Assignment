import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from time import sleep
from datetime import datetime


# Function to get the WebDriver (with the previous usage of driver setup)
def get_chromedriver(proxy=None, headless=False):
    agent = 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.2 (KHTML, like Gecko) Chrome/116.0.1216.0 Safari/537.2'

    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument(f'user-agent={agent}')

    if headless:
        chrome_options.add_argument('--headless')

    if proxy:
        chrome_options.add_argument(f'--proxy-server={proxy}')
        print(f'Browser open with proxy IP -->> {proxy.split(":")[0]}')
    else:
        print('Chrome Driver Opening...')

    web_driver = webdriver.Chrome(options=chrome_options)
    web_driver.set_window_position(325, 30)
    web_driver.implicitly_wait(10)

    return web_driver


# Function to read keywords for the current day from the Excel sheet
def get_keywords_for_day(sheet_name, excel_path):
    # Open the Excel file
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb[sheet_name]

    keywords = []

    # Iterate through rows and get the keyword (assuming they are in column B)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
        keyword = row[0].value
        if keyword:
            keywords.append(keyword)

    return keywords


# Function to write the results (longest and smallest suggestion) back to the Excel file
def write_results_to_excel(sheet_name, keyword, longest_suggestion, smallest_suggestion, excel_path):
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb[sheet_name]

    # Find the row containing the keyword and update the longest and smallest suggestion
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
        if row[0].value == keyword:
            # Assuming Longest Option is in column C and Smallest Option is in column D
            row[0].offset(0, 1).value = longest_suggestion
            row[0].offset(0, 2).value = smallest_suggestion
            break

    # Save the changes to the same file
    wb.save(excel_path)


# Main function to orchestrate the process
def main():
    # Ask the user for the path to the Excel file
    excel_path = 'reformatted_sample.xlsx'

    # Get today's day name
    current_day_name = datetime.now().strftime("%A").capitalize()  # E.g., "Monday"

    print(f'\nCurrent Day: {current_day_name}\n')
    # Get the keywords for today
    keywords = get_keywords_for_day(current_day_name, excel_path)

    # Initialize the driver using the get_driver function
    driver = get_chromedriver()

    # Perform the Google search for each keyword
    for keyword in keywords:
        # Open Google
        driver.get("https://www.google.com")

        # Locate the search box using its name attribute
        search_box = driver.find_element(By.NAME, "q")

        # Clear the search box if any previous value exists
        search_box.clear()
        print(f'Keyword: {keyword}')
        # Type the keyword into the search box
        search_box.send_keys(keyword)
        # search_box.send_keys(Keys.RETURN)
        sleep(2)

        # Wait for suggestions to appear
        WebDriverWait(driver, 10).until(
            ec.visibility_of_element_located((By.CSS_SELECTOR, "div#Alh6id ul li div.wM6W7d span"))
        )

        # Get the suggestions
        suggestions = driver.find_elements(By.CSS_SELECTOR, "div#Alh6id ul li div.wM6W7d span")

        if suggestions:
            suggestion_texts = [suggestion.text for suggestion in suggestions if suggestion.text]

            # Get the longest and smallest suggestions
            longest_suggestion = max(suggestion_texts, key=len) if suggestion_texts else "No suggestions"
            smallest_suggestion = min(suggestion_texts, key=len) if suggestion_texts else "No suggestions"
        else:
            longest_suggestion = "No suggestions"
            smallest_suggestion = "No suggestions"

        print(f'Longest suggestion: {longest_suggestion}')
        print(f'Smallest suggestion: {smallest_suggestion}')
        print()

        # Write the results to Excel
        write_results_to_excel(current_day_name, keyword, longest_suggestion, smallest_suggestion, excel_path)

        # Sleep for 2 seconds to avoid being blocked
        sleep(2)

    # Close the WebDriver after all searches are done
    driver.quit()

    print("Results saved successfully!")


if __name__ == "__main__":
    main()
